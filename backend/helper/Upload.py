from openpyxl import Workbook, load_workbook
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np 
import csv, os, sys

from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active
def excel(file):
    df = pd.read_excel(file) #for an earlier version of Excel, you may need to use the file extension of 'xls'
    print (df)
def find_tables(input_file):
    formatted_file = input_file
    
    # Look for 44 of percentiges
    # stem out until table is found
    # extract title

    # to find egdes of every table:
    # Border of where bymbers go to words,

    # GIVENS:
    # Vertical scan
    #  | [space?]
    #  | column name
    #  | [space?]
    #  | numbers (4x4 percentiges)
    #
    # Horizontal scan
    #  | [space?]
    #  | Row name
    #  | [space?]
    #  | numbers (4x4 percentiges)
    #  | Row name
    #  | [space?]
    #  | numbers (4x4 percentiges)
    #  | [repeat?]
    #
    # Finding corners (potential titles)
    # 
    wb = load_workbook(str(input_file))

    ws1 = wb.active
    ws2 = wb.create_sheet("modifiedSheet")

    start_row = 3
    start_col = 1

    for row in ws1.iter_rows(min_row=start_row):
        for cell in row:
            # print(cell.value)
            ws2.cell(row = start_row-2, column = start_col, value=cell.value) # start_row - 2 will assign the value to the same column up 2 rows
            start_col += 1 # increment the column, for use of destination sheet
        start_row += 1 # increment the column, for use of destination sheet
        start_col = 1 # reset to first column after row processing

    wb.save("modified.xlsx")